import csv
import json

from openpyxl import load_workbook

def read_excel(path, sheet):
    workbook = load_workbook(path)
    worksheet = workbook[sheet]

    data = []  # tupil - python concept

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        data.append(row)  # add the value from row to data tupil

    workbook.close()
    return data

def read_csv(path):

    data = []

    with open(path) as f:
        reader = csv.reader(f)

        next(reader)   #Skip the header

        for row in reader:
            data.append(tuple(row))

    return data

def read_json(path):

    with open(path) as f:
        return json.load(f)


